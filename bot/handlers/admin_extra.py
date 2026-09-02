"""Админка: промокоды, рекламные кампании, пробный период, способы оплаты."""
from __future__ import annotations

import re
from datetime import datetime, timedelta
from typing import Any

from openpyxl.utils import get_column_letter

from aiogram import Bot, F, Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import BufferedInputFile, CallbackQuery, Message

from .. import texts
from ..config import Tariff
from ..keyboards import (
    alerts_menu,
    campaign_detail_menu,
    csv_menu,
    sys_channel_menu,
    operators_menu,
    tariff_detail_menu,
    tariffs_admin_menu,
    user_card_menu,
    admin_menu,
    campaign_list_menu,
    pay_toggles_menu,
    promo_detail_menu,
    promo_list_menu,
    trial_settings_menu,
)
from ..remnawave import RemnaError
from ..services import (
    Runtime,
    deliver_subscription,
    subscription_kb,
    sys_channel,
    trial_config,
)
from ..utils import fmt_date, parse_iso, utcnow


async def _resolve_target(rt: Runtime, target: str) -> dict | None:
    """TG ID или логин Remnawave -> пользователь панели."""
    target = target.strip().lstrip("@")
    if target.isdigit():
        user = await rt.remna.get_user_by_telegram_id(int(target))
        if user:
            return user
    return await rt.remna.get_user_by_username(target)

router = Router(name="admin-extra")


def _is_admin(rt: Runtime, user_id: int) -> bool:
    return user_id in rt.cfg.admin_ids


def _revenue_str(revenue: dict) -> str:
    if not revenue:
        return "—"
    symbols = {"RUB": "₽", "XTR": "⭐", "USDT": " USDT"}
    return " / ".join(f"{total:g}{symbols.get(cur, f' {cur}')}" for cur, total in revenue.items())


# ══════════════════════════ ПРОМОКОДЫ ══════════════════════════


class PromoFSM(StatesGroup):
    code = State()
    days = State()
    limit = State()
    expires = State()


async def _promo_list_text(rt: Runtime) -> str:
    promos = await rt.db.list_promos()
    if not promos:
        return texts.ADMIN_PROMO.format(list=texts.ADMIN_PROMO_EMPTY)
    lines = "".join(
        texts.ADMIN_PROMO_LINE.format(
            code=p["code"], days=p["days"], used=p["used"],
            max=p["max_uses"] or "∞",
            status=texts.ADMIN_PROMO_ON if p["active"] else texts.ADMIN_PROMO_OFF,
        )
        for p in promos
    )
    return texts.ADMIN_PROMO.format(list=lines)


@router.callback_query(F.data == "adm:promo")
async def cb_promo_list(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await query.message.edit_text(
        await _promo_list_text(rt), reply_markup=promo_list_menu(await rt.db.list_promos())
    )
    await query.answer()


@router.callback_query(F.data.startswith("adm:promo:info:"))
async def cb_promo_info(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    promo = await rt.db.get_promo(int(query.data.rsplit(":", 1)[1]))
    if promo is None:
        await query.answer(texts.PROMO_NOT_FOUND, show_alert=True)
        return
    expires = "бессрочно"
    if promo["expires_at"]:
        parsed = parse_iso(promo["expires_at"])
        expires = fmt_date(parsed, rt.cfg.tz) if parsed else promo["expires_at"]
    await query.message.edit_text(
        texts.ADMIN_PROMO_DETAIL.format(
            code=promo["code"], days=promo["days"], used=promo["used"],
            max=promo["max_uses"] or "∞", expires=expires,
            status=texts.ADMIN_TRIAL_ON if promo["active"] else texts.ADMIN_TRIAL_OFF,
        ),
        reply_markup=promo_detail_menu(promo),
    )
    await query.answer()


@router.callback_query(F.data.startswith("adm:promo:tg:"))
async def cb_promo_toggle(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    promo = await rt.db.get_promo(int(query.data.rsplit(":", 1)[1]))
    if promo is None:
        await query.answer(texts.PROMO_NOT_FOUND, show_alert=True)
        return
    await rt.db.set_promo_active(promo["id"], not promo["active"])
    await cb_promo_info(query, rt)


@router.callback_query(F.data.startswith("adm:promo:del:"))
async def cb_promo_delete(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await rt.db.delete_promo(int(query.data.rsplit(":", 1)[1]))
    await query.answer(texts.ADMIN_PROMO_DELETED, show_alert=True)
    await cb_promo_list(query, rt)


@router.callback_query(F.data == "adm:promo:new")
async def cb_promo_new(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(PromoFSM.code)
    await query.message.edit_text(texts.ADMIN_PROMO_NEW_ASK_CODE)
    await query.answer()


@router.message(PromoFSM.code)
async def promo_code_input(message: Message, state: FSMContext, rt: Runtime):
    code = (message.text or "").strip().upper()
    if code == "/CANCEL":
        await state.clear()
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not re.fullmatch(r"[A-Z0-9_-]{3,32}", code):
        await message.answer("Код должен быть из латиницы/цифр, 3–32 символа. Попробуйте ещё раз:")
        return
    await state.update_data(code=code)
    await state.set_state(PromoFSM.days)
    await message.answer(texts.ADMIN_PROMO_NEW_ASK_DAYS)


@router.message(PromoFSM.days)
async def promo_days_input(message: Message, state: FSMContext, rt: Runtime):
    if not (message.text or "").strip().isdigit():
        await message.answer(texts.ADMIN_ASK_NUMBER)
        return
    days = int(message.text.strip())
    if not (1 <= days <= 3650):
        await message.answer("Дней должно быть от 1 до 3650:")
        return
    await state.update_data(days=days)
    await state.set_state(PromoFSM.limit)
    await message.answer(texts.ADMIN_PROMO_NEW_ASK_LIMIT)


@router.message(PromoFSM.limit)
async def promo_limit_input(message: Message, state: FSMContext, rt: Runtime):
    if not (message.text or "").strip().isdigit():
        await message.answer(texts.ADMIN_ASK_NUMBER)
        return
    await state.update_data(limit=int(message.text.strip()))
    await state.set_state(PromoFSM.expires)
    await message.answer(texts.ADMIN_PROMO_NEW_ASK_EXPIRES)


@router.message(PromoFSM.expires)
async def promo_expires_input(message: Message, state: FSMContext, rt: Runtime, bot: Bot):
    raw = (message.text or "").strip()
    if not raw.isdigit():
        await message.answer(texts.ADMIN_ASK_NUMBER)
        return
    data = await state.get_data()
    await state.clear()
    expires_at = None
    expires_str = "бессрочно"
    if int(raw) > 0:
        expires_at = (utcnow() + timedelta(days=int(raw))).isoformat()
        expires_str = f"через {raw} дн."

    if not await rt.db.create_promo(data["code"], data["days"], data["limit"], expires_at):
        await message.answer(texts.ADMIN_PROMO_EXISTS, reply_markup=admin_menu())
        return
    await message.answer(
        texts.ADMIN_PROMO_CREATED.format(
            code=data["code"], days=data["days"], max=data["limit"] or "∞", expires=expires_str,
        ),
        reply_markup=admin_menu(),
    )


# ══════════════════════════ РЕКЛАМНЫЕ КАМПАНИИ ══════════════════════════


class CampFSM(StatesGroup):
    name = State()


async def _camp_list_text(rt: Runtime) -> str:
    campaigns = await rt.db.list_campaigns()
    stats = await rt.db.campaign_stats()
    if not campaigns:
        return texts.ADMIN_CAMP.format(list=texts.ADMIN_CAMP_EMPTY)
    lines = ""
    for c in campaigns:
        item = stats.get(c["name"], {})
        users = item.get("users", 0)
        paid = item.get("paid", 0)
        conv = f"{paid / users * 100:.0f}%" if users else "—"
        lines += texts.ADMIN_CAMP_LINE.format(
            name=c["name"], users=users, paid=paid, conv=conv,
            revenue=_revenue_str(item.get("revenue", {})),
        )
    return texts.ADMIN_CAMP.format(list=lines)


@router.callback_query(F.data == "adm:camp")
async def cb_camp_list(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    campaigns = await rt.db.list_campaigns()
    stats = await rt.db.campaign_stats()
    await query.message.edit_text(
        await _camp_list_text(rt),
        reply_markup=campaign_list_menu(campaigns, stats),
    )
    await query.answer()


@router.callback_query(F.data.startswith("adm:camp:info:"))
async def cb_camp_info(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    cid = int(query.data.rsplit(":", 1)[1])
    campaigns = await rt.db.list_campaigns()
    camp = next((c for c in campaigns if c["id"] == cid), None)
    if camp is None:
        await query.answer("Кампания не найдена", show_alert=True)
        return
    name = camp["name"]
    d = await rt.db.campaign_detail(name)

    days_block = ""
    if d["by_day"]:
        days_block = texts.ADMIN_CAMP_DAYS_TITLE
        days_block += "".join(
            texts.ADMIN_CAMP_DAY_LINE.format(day=day, count=cnt)
            for day, cnt in d["by_day"]
        )

    icons = {"delivered": "✅", "paid": "✅", "pending": "⏳",
             "canceled": "❌", "error": "⚠️"}
    status_ru = {"delivered": "выдана", "paid": "оплачена", "pending": "ожидает",
                 "canceled": "отменена", "error": "ошибка"}
    cur_sym = {"RUB": "₽", "XTR": "⭐", "USDT": "USDT"}
    recent_block = ""
    if d["recent"]:
        recent_block = texts.ADMIN_CAMP_RECENT_TITLE
        for pay in d["recent"]:
            created = parse_iso(pay["created_at"])
            recent_block += texts.ADMIN_CAMP_RECENT_LINE.format(
                icon=icons.get(pay["status"], "•"), uid=pay["tg_id"],
                amount=f"{pay['amount']}{cur_sym.get(pay['currency'], '')}",
                status_ru=status_ru.get(pay["status"], pay["status"]),
                date=fmt_date(created, rt.cfg.tz),
            )

    await query.message.edit_text(
        texts.ADMIN_CAMP_DETAIL.format(
            name=name, users=d["users"], paid=d["paid"], conv=d["conv"],
            revenue=_revenue_str(await rt.db.campaign_stats_revenue(name)),
            days_block=days_block, recent_block=recent_block,
        ),
        reply_markup=campaign_detail_menu(cid, name, rt.bot_username),
        disable_web_page_preview=True,
    )
    await query.answer()


@router.callback_query(F.data.startswith("adm:camp:link:"))
async def cb_camp_link(query: CallbackQuery, rt: Runtime):
    """Показ ссылки текстом — можно копировать сколько угодно раз."""
    if not _is_admin(rt, query.from_user.id):
        return
    cid = int(query.data.rsplit(":", 1)[1])
    campaigns = await rt.db.list_campaigns()
    camp = next((c for c in campaigns if c["id"] == cid), None)
    if camp is None:
        await query.answer("Кампания не найдена", show_alert=True)
        return
    link = f"https://t.me/{rt.bot_username}?start=ref_{camp['name']}"
    await query.message.answer(
        f"🔗 <b>Ссылка кампании «{camp['name']}»</b> (нажмите, чтобы скопировать):\n"
        f"<code>{link}</code>\n\n"
        f"Можно копировать и использовать сколько угодно раз — все переходы "
        f"по ней учитываются в этой кампании."
    )
    await query.answer()


@router.callback_query(F.data.startswith("adm:camp:del:"))
async def cb_camp_delete(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await rt.db.delete_campaign(int(query.data.rsplit(":", 1)[1]))
    await query.answer(texts.ADMIN_CAMP_DELETED, show_alert=True)
    campaigns = await rt.db.list_campaigns()
    stats = await rt.db.campaign_stats()
    await query.message.edit_text(
        await _camp_list_text(rt),
        reply_markup=campaign_list_menu(campaigns, stats),
    )


@router.callback_query(F.data == "adm:camp:new")
async def cb_camp_new(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(CampFSM.name)
    await query.message.edit_text(texts.ADMIN_CAMP_NEW_ASK)
    await query.answer()


@router.message(CampFSM.name)
async def camp_name_input(message: Message, state: FSMContext, rt: Runtime):
    name = (message.text or "").strip().lower().lstrip("@")
    await state.clear()
    if name == "/cancel":
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not re.fullmatch(r"[a-z0-9_-]{2,32}", name):
        await message.answer(
            "Имя: латиница/цифры/дефис, 2–32 символа. Попробуйте ещё раз:",
            reply_markup=admin_menu(),
        )
        return
    if not await rt.db.add_campaign(name):
        await message.answer(texts.ADMIN_CAMP_EXISTS, reply_markup=admin_menu())
        return
    await message.answer(
        texts.ADMIN_CAMP_CREATED.format(name=name, link=f"https://t.me/{rt.bot_username}?start=ref_{name}"),
        reply_markup=admin_menu(),
        disable_web_page_preview=True,
    )


# ══════════════════════════ ПРОБНЫЙ ПЕРИОД ══════════════════════════


class SysChannelFSM(StatesGroup):
    channel = State()


class OperatorFSM(StatesGroup):
    add = State()


class TariffFSM(StatesGroup):
    new_title = State()
    new_days = State()
    new_price = State()
    new_desc = State()
    edit_value = State()


class TrialFSM(StatesGroup):
    channel = State()
    url = State()
    days = State()
    bonus_days = State()
    traffic = State()


async def _trial_text_and_kb(rt: Runtime):
    tcfg = await trial_config(rt)
    used = await rt.db.trials_count()
    channel_note = (
        tcfg["channel"] if tcfg["channel"]
        else "не задан → триал всем, БЕЗ подписки"
    )
    text = (
        "🎁 <b>Пробный период</b>\n\n"
        f"Статус: {texts.ADMIN_TRIAL_ON if tcfg['enabled'] else texts.ADMIN_TRIAL_OFF}\n"
        f"Сразу при активации: <b>{tcfg['days']} дн.</b>\n"
        f"Бонус за подписку на канал: <b>+{tcfg['bonus_days']} дн.</b>\n"
        f"Лимит трафика: <b>{tcfg['traffic_gb']} ГБ</b>\n"
        f"Канал: <code>{channel_note}</code>\n"
        f"Ссылка: <code>{tcfg['url'] or 'не задана'}</code>\n"
        f"Активировали: <b>{used}</b> чел.\n\n"
        "Схема: юзер получает базовые дни сразу. Если канал задан — может получить "
        f"ещё +{tcfg['bonus_days']} дн., подписавшись на канал (бот проверяет автоматически)."
    )
    return text, trial_settings_menu(tcfg["enabled"])


@router.callback_query(F.data == "adm:trial")
async def cb_trial_menu(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    text, kb = await _trial_text_and_kb(rt)
    await query.message.edit_text(text, reply_markup=kb)
    await query.answer()


@router.callback_query(F.data == "adm:trial:toggle")
async def cb_trial_toggle(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    tcfg = await trial_config(rt)
    await rt.db.set_setting("trial_enabled", "0" if tcfg["enabled"] else "1")
    text, kb = await _trial_text_and_kb(rt)
    await query.message.edit_text(text, reply_markup=kb)
    await query.answer("Сохранено")


@router.callback_query(F.data == "adm:trial:chan")
async def cb_trial_chan(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(TrialFSM.channel)
    await query.message.edit_text(texts.ADMIN_TRIAL_SET_CHANNEL)
    await query.answer()


@router.message(TrialFSM.channel)
async def trial_chan_input(message: Message, state: FSMContext, rt: Runtime):
    channel = (message.text or "").strip()
    await state.clear()
    if channel.lower() == "/cancel":
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not (channel.startswith("@") or channel.lstrip("-").isdigit()):
        await message.answer("Нужен @username или числовой ID. Попробуйте ещё раз:")
        return
    await rt.db.set_setting("trial_channel", channel)
    await message.answer(
        texts.ADMIN_TRIAL_SET_OK.format(key="канал", value=channel),
        reply_markup=admin_menu(),
    )


@router.callback_query(F.data == "adm:trial:url")
async def cb_trial_url(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(TrialFSM.url)
    await query.message.edit_text(texts.ADMIN_TRIAL_SET_URL)
    await query.answer()


@router.message(TrialFSM.url)
async def trial_url_input(message: Message, state: FSMContext, rt: Runtime):
    url = (message.text or "").strip()
    await state.clear()
    if url.lower() == "/cancel":
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not url.startswith("https://t.me/"):
        await message.answer("Ссылка должна начинаться с https://t.me/. Попробуйте ещё раз:")
        return
    await rt.db.set_setting("trial_channel_url", url)
    await message.answer(
        texts.ADMIN_TRIAL_SET_OK.format(key="ссылка", value=url),
        reply_markup=admin_menu(),
    )


@router.callback_query(F.data == "adm:trial:bonus")
async def cb_trial_bonus_days(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(TrialFSM.bonus_days)
    await query.message.edit_text(
        "Сколько бонусных дней давать за подписку на канал? Пришлите число (0 — выключить бонус)."
    )
    await query.answer()


@router.message(TrialFSM.bonus_days)
async def trial_bonus_days_input(message: Message, state: FSMContext, rt: Runtime):
    raw = (message.text or "").strip()
    await state.clear()
    if raw.lower() == "/cancel":
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not raw.isdigit() or int(raw) > 365:
        await message.answer("Число от 0 до 365. Попробуйте ещё раз:")
        return
    await rt.db.set_setting("trial_bonus_days", raw)
    await message.answer(
        texts.ADMIN_TRIAL_SET_OK.format(key="бонус за подписку", value=f"+{raw} дн."),
        reply_markup=admin_menu(),
    )


@router.callback_query(F.data == "adm:trial:traffic")
async def cb_trial_traffic(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(TrialFSM.traffic)
    await query.message.edit_text(
        "Лимит трафика пробного периода в ГБ. Пришлите число (например 15)."
    )
    await query.answer()


@router.message(TrialFSM.traffic)
async def trial_traffic_input(message: Message, state: FSMContext, rt: Runtime):
    raw = (message.text or "").strip()
    await state.clear()
    if raw.lower() == "/cancel":
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not raw.isdigit() or not (1 <= int(raw) <= 10000):
        await message.answer("Число ГБ от 1 до 10000. Попробуйте ещё раз:")
        return
    await rt.db.set_setting("trial_traffic_gb", raw)
    await message.answer(
        texts.ADMIN_TRIAL_SET_OK.format(key="лимит трафика", value=f"{raw} ГБ"),
        reply_markup=admin_menu(),
    )


@router.callback_query(F.data == "adm:trial:days")
async def cb_trial_days(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(TrialFSM.days)
    await query.message.edit_text(texts.ADMIN_TRIAL_SET_DAYS)
    await query.answer()


@router.message(TrialFSM.days)
async def trial_days_input(message: Message, state: FSMContext, rt: Runtime):
    raw = (message.text or "").strip()
    await state.clear()
    if raw.lower() == "/cancel":
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not raw.isdigit() or not (1 <= int(raw) <= 365):
        await message.answer("Число от 1 до 365. Попробуйте ещё раз:")
        return
    await rt.db.set_setting("trial_days", raw)
    await message.answer(
        texts.ADMIN_TRIAL_SET_OK.format(key="срок", value=f"{raw} дн."),
        reply_markup=admin_menu(),
    )


# ══════════════════════════ СПОСОБЫ ОПЛАТЫ ══════════════════════════


async def _pay_text_and_kb(rt: Runtime):
    configured = {
        "card": bool(await rt.db.get_setting("card_number", "")),
        "yookassa": rt.yookassa is not None and rt.cfg.yookassa_enabled,
        "stars": rt.cfg.stars_enabled,
        "cryptobot": rt.cryptobot is not None and rt.cfg.cryptobot_enabled,
    }
    defaults = {"card": "1", "yookassa": "0", "stars": "0", "cryptobot": "0"}
    states = {
        name: await rt.db.get_setting(f"pay_{name}", defaults[name]) == "1"
        for name in configured
    }
    labels = {"card": "Перевод на карту (ручное подтверждение)", "yookassa": "ЮKassa (карты)",
              "stars": "Telegram Stars", "cryptobot": "CryptoBot (USDT)"}
    lines = "".join(
        texts.ADMIN_PAY_LINE.format(
            icon="✅" if states[name] else "❌", label=labels[name],
        )
        for name, conf in configured.items() if conf
    ) or "ни один способ не настроен"
    return texts.ADMIN_PAY.format(list=lines), pay_toggles_menu(states, configured)


@router.callback_query(F.data == "adm:pay")
async def cb_pay_menu(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    text, kb = await _pay_text_and_kb(rt)
    await query.message.edit_text(text, reply_markup=kb)
    await query.answer()


@router.callback_query(F.data.startswith("adm:pay:"))
async def cb_pay_toggle(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    name = query.data.rsplit(":", 1)[1]
    if name not in ("stars", "cryptobot", "yookassa"):
        await query.answer()
        return
    current = await rt.db.get_setting(f"pay_{name}", "1") == "1"
    await rt.db.set_setting(f"pay_{name}", "0" if current else "1")
    text, kb = await _pay_text_and_kb(rt)
    await query.message.edit_text(text, reply_markup=kb)
    await query.answer("Сохранено")


# ══════════════════════════ КАРТОЧКА ПОЛЬЗОВАТЕЛЯ ══════════════════════════


class UserCardFSM(StatesGroup):
    target = State()
    extend_days = State()


async def _render_user_card(rt: Runtime, rw_user: dict | None, target: str) -> tuple[str, Any]:
    """Собирает карточку. rw_user может быть None (нет в Remnawave)."""
    tg_id = None
    bot_user = None
    if target.isdigit():
        tg_id = int(target)
        bot_user = await rt.db.get_bot_user(tg_id)
    elif rw_user and rw_user.get("telegramId"):
        try:
            tg_id = int(rw_user["telegramId"])
        except (TypeError, ValueError):
            tg_id = None
        bot_user = await rt.db.get_bot_user(tg_id)

    payments = await rt.db.payments_for_user(tg_id, limit=5) if tg_id else []
    if payments:
        lines = ""
        cur_sym = {"RUB": "₽", "XTR": "⭐", "USDT": " USDT"}
        status_ru = {"delivered": "✅", "paid": "✅", "pending": "⏳",
                     "canceled": "❌", "error": "⚠️"}
        for pay in payments:
            created = parse_iso(pay["created_at"])
            lines += (
                f"├ #{pay['id']} {fmt_date(created, rt.cfg.tz)} — "
                f"{pay['amount']}{cur_sym.get(pay['currency'], '')} "
                f"({pay['provider']}) {status_ru.get(pay['status'], pay['status'])}\n"
            )
    else:
        lines = "├ " + texts.ADMIN_USER_NO_PAYMENTS + "\n"

    source = "— напрямую"
    if bot_user:
        if bot_user.get("source"):
            source = f"реклама «{bot_user['source']}»"
        elif bot_user.get("referred_by"):
            source = f"реферал <code>{bot_user['referred_by']}</code>"

    referrals = await rt.db.count_referrals(tg_id) if tg_id else 0
    paid_ref = await rt.db.paid_referrals(tg_id) if tg_id else 0

    if rw_user:
        from ..utils import fmt_bytes

        expire = parse_iso(rw_user.get("expireAt"))
        used = rw_user.get("usedTrafficBytes")
        if used is None:
            used = (rw_user.get("userTraffic") or {}).get("usedTrafficBytes")
        traffic = fmt_bytes(used)
        limit_b = rw_user.get("trafficLimitBytes")
        if limit_b:
            traffic += " / " + fmt_bytes(limit_b)
        rw_block = {
            "rw_username": rw_user.get("username", "—"),
            "rw_status": rw_user.get("status", "—"),
            "expire": fmt_date(expire, rt.cfg.tz),
            "traffic": traffic,
        }
    else:
        rw_block = {
            "rw_username": "— (нет в Remnawave)",
            "rw_status": "—",
            "expire": "—",
            "traffic": "—",
        }

    if tg_id:
        name = bot_user.get("first_name") if bot_user else None
        uname = bot_user.get("username") if bot_user else None
        tg_line = (
            f"<a href=\"tg://user?id={tg_id}\">{name or uname or tg_id}</a> (<code>{tg_id}</code>)"
        )
    else:
        tg_line = texts.ADMIN_USER_NOT_IN_BOT

    text = texts.ADMIN_USER_CARD.format(
        tg_line=tg_line,
        referrals=referrals,
        paid_ref=paid_ref,
        trial=("использован" if bot_user and bot_user.get("trial_used") else "нет"),
        payments=lines,
        source=source,
        **rw_block,
    )
    disabled = bool(rw_user and str(rw_user.get("status", "")).upper() == "DISABLED")
    kb = user_card_menu(tg_id, rw_user.get("uuid") if rw_user else None, disabled)
    return text, kb


@router.callback_query(F.data == "adm:user")
async def cb_user_card(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(UserCardFSM.target)
    await query.message.edit_text(texts.ADMIN_USER_ASK)
    await query.answer()


@router.message(UserCardFSM.target)
async def user_card_input(message: Message, state: FSMContext, rt: Runtime):
    target = (message.text or "").strip().lstrip("@")
    if target.lower() == "/cancel":
        await state.clear()
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not target:
        return
    rw_user = await _resolve_target(rt, target)
    if rw_user is None and not target.isdigit():
        await message.answer(texts.ADMIN_USER_NOT_FOUND)
        return
    await state.clear()
    await state.update_data(uc_target=target)
    text, kb = await _render_user_card(rt, rw_user, target)
    await message.answer(text, reply_markup=kb, disable_web_page_preview=True)


@router.callback_query(F.data.startswith("adm:uc:toggle:"))
async def cb_user_card_toggle(query: CallbackQuery, rt: Runtime, bot: Bot):
    if not _is_admin(rt, query.from_user.id):
        return
    uuid = query.data.rsplit(":", 1)[1]
    try:
        rw_user = await rt.remna.get_user(uuid)
    except RemnaError as e:
        await query.answer(f"Ошибка: {e}", show_alert=True)
        return
    disable = str(rw_user.get("status", "")).upper() != "DISABLED"
    try:
        if disable:
            await rt.remna.disable_user(uuid)
        else:
            await rt.remna.enable_user(uuid)
    except RemnaError as e:
        await query.answer(f"Ошибка: {e}", show_alert=True)
        return
    await query.answer("Готово")
    text, kb = await _render_user_card(rt, rw_user, str(rw_user.get("telegramId") or uuid))
    try:
        await query.message.edit_text(text, reply_markup=kb, disable_web_page_preview=True)
    except Exception:
        pass


@router.callback_query(F.data == "adm:uc:extend")
async def cb_user_card_extend(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(UserCardFSM.extend_days)
    await query.message.answer(texts.ADMIN_USER_EXTEND_ASK)
    await query.answer()


@router.message(UserCardFSM.extend_days)
async def user_card_extend_days(message: Message, state: FSMContext, rt: Runtime, bot: Bot):
    raw = (message.text or "").strip()
    if not raw.isdigit():
        await message.answer(texts.ADMIN_ASK_NUMBER)
        return
    days = int(raw)
    data = await state.get_data()
    await state.clear()
    target = data.get("uc_target")
    if not target:
        await message.answer("Пользователь не выбран. Начните заново: /admin")
        return
    rw_user = await _resolve_target(rt, target)
    tariff = Tariff(id="gift", title=f"Продление админом ({days} дн.)", days=days, description="")
    try:
        result_text, url = await deliver_subscription(
            rt, int(target) if target.isdigit() else None, tariff, existing=rw_user
        )
    except RemnaError as e:
        await message.answer(f"❌ Ошибка: {e}")
        return
    await rt.db.add_payment(
        int(target) if target.isdigit() else 0, "gift", "admin", "0", "-",
        status="delivered", note=f"card extend {days}d target={target}",
    )
    await message.answer(texts.ADMIN_GRANT_DONE.format(details=result_text),
                         disable_web_page_preview=True)
    notify_tg = int(target) if target.isdigit() else None
    if notify_tg is None and rw_user and rw_user.get("telegramId"):
        try:
            notify_tg = int(rw_user["telegramId"])
        except (TypeError, ValueError):
            notify_tg = None
    if notify_tg:
        try:
            await bot.send_message(
                notify_tg, "🎁 Администратор продлил вам подписку!\n\n" + result_text,
                reply_markup=subscription_kb(url), disable_web_page_preview=True,
            )
        except Exception:
            pass


# ══════════════════════════ CSV-ЭКСПОРТ ══════════════════════════


@router.callback_query(F.data == "adm:csv")
async def cb_csv_menu(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await query.message.edit_text(texts.ADMIN_CSV_MENU, reply_markup=csv_menu())
    await query.answer()


def _csv_response(filename: str, header: list[str], rows: list[list]) -> BufferedInputFile:
    import csv as csvlib
    import io

    buf = io.StringIO()
    writer = csvlib.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    return BufferedInputFile(buf.getvalue().encode("utf-8-sig"), filename=filename)


@router.callback_query(F.data == "adm:csv:payments")
async def cb_csv_payments(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await query.answer("Готовлю файл…")
    rows = [
        [p["id"], p["created_at"], p["tg_id"], p["tariff_id"], p["provider"],
         p["amount"], p["currency"], p["status"], p["source"] or ""]
        for p in await rt.db.all_payments()
    ]
    file = _csv_response(
        f"payments_{datetime.now():%Y%m%d}.csv",
        ["id", "created_at", "tg_id", "tariff", "provider", "amount", "currency",
         "status", "source"],
        rows,
    )
    await query.message.answer_document(file, caption=f"🧾 Платежи: {len(rows)}")


@router.callback_query(F.data == "adm:csv:users")
async def cb_csv_users(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await query.answer("Готовлю файл…")
    rows = [
        [u["tg_id"], u["username"] or "", u["created_at"], u["source"] or "",
         u["referred_by"] or "", "yes" if u["trial_used"] else "no"]
        for u in await rt.db.all_bot_users_full()
    ]
    file = _csv_response(
        f"users_{datetime.now():%Y%m%d}.csv",
        ["tg_id", "username", "created_at", "source", "referred_by", "trial_used"],
        rows,
    )
    await query.message.answer_document(file, caption=f"👥 Пользователи: {len(rows)}")


# ══════════════════════════ ОТЧЁТЫ И АЛЕРТЫ ══════════════════════════


async def _alerts_text_and_kb(rt: Runtime):
    reports = await rt.db.get_setting("reports_enabled", "1") == "1"
    nodes = await rt.db.get_setting("alerts_nodes", "1") == "1"
    backup = await rt.db.get_setting("alerts_backup", "1") == "1"
    text = (
        "🔔 <b>Отчёты и алерты</b>\n\n"
        "📊 Отчёт админу каждые 4 часа: продажи, новые юзеры, статус нод\n"
        "🔴 Мгновенные алерты, когда нода уходит в оффлайн и возвращается\n"
        "💾 Ежедневный бэкап БД файлом в этот чат\n\n"
        "Нажмите на пункт, чтобы включить/выключить:"
    )
    return text, alerts_menu(
        reports_enabled=reports, node_alerts_enabled=nodes,
        backup_enabled=backup, interval_h=4,
    )


@router.callback_query(F.data == "adm:alerts")
async def cb_alerts(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    text, kb = await _alerts_text_and_kb(rt)
    await query.message.edit_text(text, reply_markup=kb)
    await query.answer()


@router.callback_query(F.data.startswith("adm:al:"))
async def cb_alerts_toggle(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    key = query.data.rsplit(":", 1)[1]
    setting = {"reports": "reports_enabled", "nodes": "alerts_nodes",
               "backup": "alerts_backup"}.get(key)
    if not setting:
        await query.answer()
        return
    current = await rt.db.get_setting(setting, "1") == "1"
    await rt.db.set_setting(setting, "0" if current else "1")
    text, kb = await _alerts_text_and_kb(rt)
    await query.message.edit_text(text, reply_markup=kb)
    await query.answer("Сохранено")


# ══════════════════════════ EXCEL-ЭКСПОРТ ══════════════════════════


async def _xlsx_data(rt: Runtime) -> dict:
    """Собирает все данные для книги Excel (асинхронная часть)."""
    return {
        "stats": await rt.db.sales_stats(),
        "payments": await rt.db.all_payments(),
        "users": await rt.db.all_bot_users_full(),
        "campaigns": await rt.db.campaign_stats(),
        "trials": await rt.db.trials_count(),
    }


def _build_xlsx(rt: Runtime, data: dict) -> BufferedInputFile:
    """Книга из 4 листов: Сводка с графиком, Платежи, Пользователи, Реклама."""
    import io
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill

    cur_sym = {"RUB": "₽", "XTR": "⭐", "USDT": " USDT"}
    status_ru = {"delivered": "выдана", "paid": "оплачена", "pending": "ожидает",
                 "canceled": "отменена", "error": "ошибка"}
    provider_ru = {"card": "Перевод на карту", "lava": "Lava", "stars": "Telegram Stars",
                   "cryptobot": "CryptoBot", "yookassa": "ЮKassa",
                   "promo": "Промокод", "trial": "Пробный период",
                   "admin": "Админ", "refbonus": "Реф. бонус"}

    bold = Font(bold=True)
    h2 = Font(bold=True, size=13)
    fill = PatternFill("solid", fgColor="D9E1F2")

    wb = Workbook()

    # ── Лист 1: Сводка ──
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Сводка магазина"
    ws["A1"].font = h2
    st = data["stats"]
    rows = [
        ("Пользователей бота", st["bot_users"]),
        ("Оплат всего", sum(r["cnt"] for r in st["by_provider"])),
        ("Оплат за 7 дней", st["week"]),
        ("Оплат за 30 дней", st["month"]),
        ("Выдано вручную", st["gifts"]),
        ("Пробных активаций", data.get("trials", 0)),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=value)

    # продажи по провайдерам (только рублёвые суммы считаем деньгами)
    start = 10
    ws.cell(row=start, column=1, value="Продажи по способам оплаты").font = bold
    ws.cell(row=start + 1, column=1, value="Способ").fill = fill
    ws.cell(row=start + 1, column=2, value="Оплат").fill = fill
    ws.cell(row=start + 1, column=3, value="Сумма").fill = fill
    r = start + 2
    for row in st["by_provider"]:
        name = provider_ru.get(row["provider"], row["provider"])
        cur = row["currency"]
        total = row["total"] or 0
        ws.cell(row=r, column=1, value=f"{name} ({cur})")
        ws.cell(row=r, column=2, value=row["cnt"])
        ws.cell(row=r, column=3, value=total)
        r += 1

    chart = BarChart()
    chart.title = "Оплаты по способам"
    chart.y_axis.title = "Число оплат"
    data_ref = Reference(ws, min_col=2, min_row=start + 1, max_row=r - 1)
    cats = Reference(ws, min_col=1, min_row=start + 2, max_row=r - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, f"E{start + 1}")

    # выручка по дням (30 дней) — вторая таблица + график
    daily = data.get("daily") or []
    start2 = r + 2
    ws.cell(row=start2, column=1, value="Выручка по дням (30 дн.), ₽").font = bold
    ws.cell(row=start2 + 1, column=1, value="День").fill = fill
    ws.cell(row=start2 + 1, column=2, value="Оплат").fill = fill
    ws.cell(row=start2 + 1, column=3, value="₽").fill = fill
    for i, d in enumerate(daily, start=start2 + 2):
        ws.cell(row=i, column=1, value=d["day"])
        ws.cell(row=i, column=2, value=d["cnt"])
        ws.cell(row=i, column=3, value=round(d["rub"], 2))
    if daily:
        chart2 = BarChart()
        chart2.title = "Выручка по дням"
        dr = Reference(ws, min_col=3, min_row=start2 + 1, max_row=start2 + 1 + len(daily))
        dcats = Reference(ws, min_col=1, min_row=start2 + 2, max_row=start2 + 1 + len(daily))
        chart2.add_data(dr, titles_from_data=True)
        chart2.set_categories(dcats)
        chart2.height = 8
        chart2.width = 20
        ws.add_chart(chart2, f"E{start2 + 1}")

    # ── Лист 2: Платежи ──
    ws2 = wb.create_sheet("Платежи")
    headers = ["ID", "Дата", "TG ID", "Тариф", "Способ", "Сумма", "Статус", "Источник"]
    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = fill
    for i, p_ in enumerate(data["payments"], start=2):
        tariff = rt.cfg.tariffs.get(p_["tariff_id"])
        created = parse_iso(p_["created_at"])
        ws2.cell(row=i, column=1, value=p_["id"])
        c2 = ws2.cell(row=i, column=2, value=fmt_date(created, rt.cfg.tz))
        c2.number_format = "@"
        ws2.cell(row=i, column=3, value=p_["tg_id"])
        ws2.cell(row=i, column=4, value=tariff.title if tariff else p_["tariff_id"])
        ws2.cell(row=i, column=5, value=provider_ru.get(p_["provider"], p_["provider"]))
        ws2.cell(row=i, column=6, value=float(p_["amount"]) if p_["amount"] else 0)
        ws2.cell(row=i, column=7, value=status_ru.get(p_["status"], p_["status"]))
        ws2.cell(row=i, column=8, value=p_["source"] or "")
    widths = [6, 18, 14, 24, 20, 10, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── Лист 3: Пользователи ──
    ws3 = wb.create_sheet("Пользователи")
    headers3 = ["TG ID", "Имя", "Username", "Регистрация", "Источник", "Реферал", "Пробный"]
    for c, h in enumerate(headers3, start=1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = fill
    for i, u in enumerate(data["users"], start=2):
        created = parse_iso(u["created_at"])
        ws3.cell(row=i, column=1, value=u["tg_id"])
        ws3.cell(row=i, column=2, value=u["first_name"] or "")
        ws3.cell(row=i, column=3, value=u["username"] or "")
        ws3.cell(row=i, column=4, value=fmt_date(created, rt.cfg.tz))
        ws3.cell(row=i, column=5, value=u["source"] or "")
        ws3.cell(row=i, column=6, value=u["referred_by"] or "")
        ws3.cell(row=i, column=7, value="да" if u["trial_used"] else "")
    for i, w in enumerate([14, 18, 16, 18, 16, 14, 10], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # ── Лист 4: Реклама ──
    ws4 = wb.create_sheet("Реклама")
    headers4 = ["Кампания", "Юзеров", "Оплат", "Выручка"]
    for c, h in enumerate(headers4, start=1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = fill
    for i, (name, item) in enumerate(data["campaigns"].items(), start=2):
        revenue = " / ".join(
            f"{total:g}{cur_sym.get(cur, cur)}" for cur, total in item["revenue"].items()
        ) or "—"
        ws4.cell(row=i, column=1, value=name)
        ws4.cell(row=i, column=2, value=item["users"])
        ws4.cell(row=i, column=3, value=item["paid"])
        ws4.cell(row=i, column=4, value=revenue)
    for i, w in enumerate([20, 10, 10, 18], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    name = f"report_{datetime.now():%Y%m%d_%H%M}.xlsx"
    return BufferedInputFile(buf.getvalue(), filename=name)


async def _trials(rt: Runtime) -> int:
    return await rt.db.trials_count()


async def _daily_revenue(rt: Runtime, days: int = 30) -> list[dict]:
    rows = []
    for d in range(days - 1, -1, -1):
        day_start = (utcnow() - timedelta(days=d)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        rows.append({
            "day": day_start.strftime("%d.%m"),
            "cnt": await rt.db.sales_count_between(
                day_start.isoformat(), day_end.isoformat()),
            "rub": await rt.db.revenue_rub_between(
                day_start.isoformat(), day_end.isoformat()),
        })
    return rows


@router.callback_query(F.data == "adm:csv:xlsx")
async def cb_xlsx(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await query.answer("Готовлю книгу Excel…")
    data = await _xlsx_data(rt)
    data["daily"] = await _daily_revenue(rt)
    file = _build_xlsx(rt, data)
    await query.message.answer_document(file, caption="📊 Полный отчёт магазина (Excel)")


# ══════════════════════════ ОПЕРАТОРЫ ОПЛАТЫ ══════════════════════════


async def _operators_text_and_kb(rt: Runtime):
    ops = await rt.db.payment_operators()
    listing = "\n".join(f"├ 👤 <code>{tg}</code>" for tg in ops) or (
        "├ " + texts.ADMIN_OPERATORS_EMPTY
    )
    return texts.ADMIN_OPERATORS.format(list=listing + "\n"), operators_menu(ops)


@router.callback_query(F.data == "adm:operators")
async def cb_operators(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    text, kb = await _operators_text_and_kb(rt)
    await query.message.edit_text(text, reply_markup=kb)
    await query.answer()


@router.callback_query(F.data == "adm:op:add")
async def cb_operator_add(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(OperatorFSM.add)
    await query.message.edit_text(texts.ADMIN_OPERATORS_ADD_ASK)
    await query.answer()


@router.message(OperatorFSM.add)
async def operator_add_input(message: Message, state: FSMContext, rt: Runtime):
    value = (message.text or "").strip()
    await state.clear()
    if value.lower() == "/cancel":
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not value.lstrip("-").isdigit():
        await message.answer("Нужен числовой Telegram ID. Попробуйте ещё раз:")
        await state.set_state(OperatorFSM.add)
        return
    tg_id = int(value)
    ops = await rt.db.payment_operators()
    if tg_id in rt.cfg.admin_ids:
        return await message.answer("Это администратор — он и так получает чеки.",
                                    reply_markup=admin_menu())
    if tg_id in ops:
        return await message.answer("Этот оператор уже добавлен.", reply_markup=admin_menu())
    ops.append(tg_id)
    await rt.db.set_setting("pay_operators", ",".join(map(str, ops)))
    await message.answer(
        texts.ADMIN_OPERATORS_ADDED.format(id=tg_id), reply_markup=admin_menu()
    )


@router.callback_query(F.data.startswith("adm:op:del:"))
async def cb_operator_del(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    tg_id = int(query.data.rsplit(":", 1)[1])
    ops = await rt.db.payment_operators()
    if tg_id in ops:
        ops.remove(tg_id)
        await rt.db.set_setting("pay_operators", ",".join(map(str, ops)))
    await query.answer(texts.ADMIN_OPERATORS_REMOVED.format(id=tg_id), show_alert=True)
    text, kb = await _operators_text_and_kb(rt)
    try:
        await query.message.edit_text(text, reply_markup=kb)
    except Exception:
        pass


# ══════════════════════════ РЕДАКТОР ТАРИФОВ ══════════════════════════


def _price_str(t) -> str:
    return f"{t.price_rub:g} ₽" if t.price_rub is not None else "без цены"


async def _tariffs_admin_text(rt: Runtime) -> str:
    tariffs = sorted(rt.cfg.tariffs.values(), key=lambda t: (not t.visible, t.days))
    if not tariffs:
        return texts.ADMIN_TARIFFS.format(list=texts.ADMIN_TARIFFS_EMPTY)
    lines = "".join(
        texts.ADMIN_TARIFF_LINE.format(
            visible=texts.ADMIN_TARIFF_ON if t.visible else texts.ADMIN_TARIFF_OFF,
            title=t.title, days=t.days, price=_price_str(t), id=t.id,
        )
        for t in tariffs
    )
    return texts.ADMIN_TARIFFS.format(list=lines)


@router.callback_query(F.data == "adm:tariffs")
async def cb_tariffs(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await query.message.edit_text(
        await _tariffs_admin_text(rt), reply_markup=tariffs_admin_menu(
            sorted(rt.cfg.tariffs.values(), key=lambda t: (not t.visible, t.days))
        )
    )
    await query.answer()


@router.callback_query(F.data.startswith("adm:tar:info:"))
async def cb_tariff_info(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    tid = query.data.rsplit(":", 1)[1]
    t = rt.cfg.tariffs.get(tid)
    if t is None:
        await query.answer("Тариф не найден", show_alert=True)
        return
    await query.message.edit_text(
        texts.ADMIN_TARIFF_DETAIL.format(
            title=t.title, id=t.id, days=t.days, price=_price_str(t),
            description=t.description or "—",
            visible=(texts.ADMIN_TARIFF_VISIBLE_ON if t.visible
                     else texts.ADMIN_TARIFF_VISIBLE_OFF),
        ),
        reply_markup=tariff_detail_menu(t),
    )
    await query.answer()


async def _save_tariffs(rt: Runtime) -> None:
    await rt.db.set_setting("tariffs_json", await rt.serialize_tariffs())
    await rt.reload_tariffs()


@router.callback_query(F.data.startswith("adm:tar:vis:"))
async def cb_tariff_visible(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    tid = query.data.rsplit(":", 1)[1]
    t = rt.cfg.tariffs.get(tid)
    if t is None:
        await query.answer("Тариф не найден", show_alert=True)
        return
    from ..config import Tariff

    rt.cfg.tariffs[tid] = Tariff(
        id=t.id, title=t.title, days=t.days, description=t.description,
        price_rub=t.price_rub, price_stars=t.price_stars, price_usdt=t.price_usdt,
        visible=not t.visible,
    )
    await _save_tariffs(rt)
    await cb_tariff_info(query, rt)


@router.callback_query(F.data.startswith("adm:tar:del:"))
async def cb_tariff_delete(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    tid = query.data.rsplit(":", 1)[1]
    if len(rt.cfg.tariffs) <= 1:
        return await query.answer(texts.ADMIN_TARIFF_LAST, show_alert=True)
    if tid in rt.cfg.tariffs:
        del rt.cfg.tariffs[tid]
        await _save_tariffs(rt)
    await query.answer(texts.ADMIN_TARIFF_DELETED, show_alert=True)
    await cb_tariffs(query, rt)


@router.callback_query(F.data.startswith("adm:tar:edit:"))
async def cb_tariff_edit(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    parts = query.data.split(":")
    tid, field = parts[3], parts[4]
    if tid not in rt.cfg.tariffs or field not in texts.ADMIN_TARIFF_EDIT_ASK:
        await query.answer("Не найдено", show_alert=True)
        return
    await state.set_state(TariffFSM.edit_value)
    await state.update_data(tid=tid, field=field)
    await query.message.answer(texts.ADMIN_TARIFF_EDIT_ASK[field])
    await query.answer()


@router.message(TariffFSM.edit_value)
async def tariff_edit_input(message: Message, state: FSMContext, rt: Runtime):
    from ..config import Tariff

    value = (message.text or "").strip()
    data = await state.get_data()
    await state.clear()
    tid, field = data.get("tid"), data.get("field")
    t = rt.cfg.tariffs.get(tid)
    if t is None or not field or value.lower() == "/cancel":
        return await message.answer("Отменено.")
    kwargs = {}
    if field == "title":
        if not (1 <= len(value) <= 64):
            await message.answer("Название 1–64 символа. Ещё раз:")
            await state.set_state(TariffFSM.edit_value)
            return
        kwargs["title"] = value
        shown = value
    elif field == "days":
        if not value.isdigit() or not (1 <= int(value) <= 3650):
            await message.answer("Число дней 1–3650. Ещё раз:")
            await state.set_state(TariffFSM.edit_value)
            return
        kwargs["days"] = int(value)
        shown = f"{value} дн."
    elif field == "price_rub":
        try:
            price = round(float(value.replace(",", ".")), 2)
            if price < 0:
                raise ValueError
        except ValueError:
            await message.answer("Цена — число (например 199). Ещё раз:")
            await state.set_state(TariffFSM.edit_value)
            return
        kwargs["price_rub"] = price or None
        shown = f"{price:g} ₽"
    elif field == "description":
        kwargs["description"] = "" if value == "-" else value[:200]
        shown = kwargs["description"] or "—"
    else:
        return await message.answer("Неизвестное поле.")
    rt.cfg.tariffs[tid] = Tariff(
        id=t.id, title=t.title, days=t.days,
        description=kwargs.get("description", t.description),
        price_rub=kwargs.get("price_rub", t.price_rub),
        price_stars=t.price_stars, price_usdt=t.price_usdt,
        visible=t.visible,
        **({"title": kwargs["title"]} if "title" in kwargs else {}),
        **({"days": kwargs["days"]} if "days" in kwargs else {}),
    )
    await _save_tariffs(rt)
    await message.answer(
        texts.ADMIN_TARIFF_UPDATED.format(
            field=texts.ADMIN_TARIFF_EDIT_FIELD_RU.get(field, field), value=shown,
        ),
        reply_markup=admin_menu(),
    )


@router.callback_query(F.data == "adm:tar:new")
async def cb_tariff_new(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(TariffFSM.new_title)
    await query.message.edit_text(texts.ADMIN_TARIFF_ASK_TITLE)
    await query.answer()


@router.message(TariffFSM.new_title)
async def tariff_new_title(message: Message, state: FSMContext, rt: Runtime):
    value = (message.text or "").strip()
    if value.lower() == "/cancel":
        await state.clear()
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not (1 <= len(value) <= 64):
        await message.answer("Название 1–64 символа:")
        return
    await state.update_data(title=value)
    await state.set_state(TariffFSM.new_days)
    await message.answer(texts.ADMIN_TARIFF_ASK_DAYS)


@router.message(TariffFSM.new_days)
async def tariff_new_days(message: Message, state: FSMContext, rt: Runtime):
    value = (message.text or "").strip()
    if not value.isdigit() or not (1 <= int(value) <= 3650):
        await message.answer(texts.ADMIN_TARIFF_ASK_DAYS)
        return
    await state.update_data(days=int(value))
    await state.set_state(TariffFSM.new_price)
    await message.answer(texts.ADMIN_TARIFF_ASK_PRICE)


@router.message(TariffFSM.new_price)
async def tariff_new_price(message: Message, state: FSMContext, rt: Runtime):
    value = (message.text or "").strip().replace(",", ".")
    try:
        price = round(float(value), 2)
        if price < 0:
            raise ValueError
    except ValueError:
        await message.answer(texts.ADMIN_TARIFF_ASK_PRICE)
        return
    await state.update_data(price_rub=price)
    await state.set_state(TariffFSM.new_desc)
    await message.answer(texts.ADMIN_TARIFF_ASK_DESC)


@router.message(TariffFSM.new_desc)
async def tariff_new_desc(message: Message, state: FSMContext, rt: Runtime):
    from ..config import Tariff

    value = (message.text or "").strip()
    await state.clear()
    if value.lower() == "/cancel":
        return await message.answer("Отменено.", reply_markup=admin_menu())
    data = await state.get_data() or {}
    title = data.get("title", "Тариф")
    days = data.get("days", 30)
    price = data.get("price_rub") or None

    # уникальный id
    base = re.sub(r"[^a-z0-9_-]", "", title.lower().replace(" ", "-"))[:20] or "tariff"
    tid = base
    n = 2
    while tid in rt.cfg.tariffs:
        tid = f"{base}{n}"
        n += 1
    rt.cfg.tariffs[tid] = Tariff(
        id=tid, title=title, days=days, description="" if value == "-" else value[:200],
        price_rub=price, visible=True,
    )
    await _save_tariffs(rt)
    await message.answer(
        texts.ADMIN_TARIFF_CREATED.format(title=title), reply_markup=admin_menu()
    )


# ══════════════════════════ СИСТЕМНЫЙ КАНАЛ ══════════════════════════


async def _sysch_text_and_kb(rt: Runtime):
    channel = await sys_channel(rt)
    current = (
        texts.ADMIN_SYS_CHANNEL_CURRENT.format(channel=channel)
        if channel else texts.ADMIN_SYS_CHANNEL_NOT_SET
    )
    text = (
        "📢 <b>Системный канал</b>\n\n"
        f"{current}\n\n"
        "В канал дублируются события с хэштегами для поиска:\n"
        "#пользователь #оплата #покупка #промокод #пробный #реферал\n"
        "#админ #тикет #нода #бэкап #отчёт #ошибка\n\n"
        "Бот должен быть администратором канала с правом публикации."
    )
    return text, sys_channel_menu(has_channel=bool(channel))


@router.callback_query(F.data == "adm:sysch")
async def cb_sysch(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    text, kb = await _sysch_text_and_kb(rt)
    await query.message.edit_text(text, reply_markup=kb)
    await query.answer()


@router.callback_query(F.data == "adm:sysch:set")
async def cb_sysch_set(query: CallbackQuery, state: FSMContext, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await state.set_state(SysChannelFSM.channel)
    await query.message.edit_text(texts.ADMIN_SYS_CHANNEL_ASK)
    await query.answer()


@router.message(SysChannelFSM.channel)
async def sysch_input(message: Message, state: FSMContext, rt: Runtime, bot: Bot):
    value = (message.text or "").strip()
    await state.clear()
    if value.lower() == "/cancel":
        return await message.answer("Отменено.", reply_markup=admin_menu())
    if not (value.startswith("@") or value.lstrip("-").isdigit()):
        await message.answer("Нужен ID вида -100... или @username. Попробуйте ещё раз:")
        await state.set_state(SysChannelFSM.channel)
        return
    chat_id = int(value) if value.lstrip("-").isdigit() else value
    try:
        await bot.send_message(chat_id, "✅ Бот подключён к этому каналу для системных событий.")
    except Exception as e:
        return await message.answer(
            f"❌ Не удалось отправить сообщение в канал ({e}).\n"
            "Добавьте бота администратором и попробуйте снова.",
            reply_markup=admin_menu(),
        )
    await rt.db.set_setting("sys_channel", value)
    await message.answer(
        texts.ADMIN_SYS_CHANNEL_SET_OK.format(channel=value), reply_markup=admin_menu()
    )


@router.callback_query(F.data == "adm:sysch:test")
async def cb_sysch_test(query: CallbackQuery, rt: Runtime, bot: Bot):
    if not _is_admin(rt, query.from_user.id):
        return
    channel = await sys_channel(rt)
    if not channel:
        return await query.answer("Канал не задан", show_alert=True)
    try:
        await bot.send_message(channel, "🧪 Тест публикации системного канала #тест")
        await query.answer(texts.ADMIN_SYS_TEST_SENT, show_alert=True)
    except Exception as e:
        await query.answer(texts.ADMIN_SYS_TEST_FAIL.format(error=str(e)[:200]),
                           show_alert=True)


@router.callback_query(F.data == "adm:sysch:del")
async def cb_sysch_del(query: CallbackQuery, rt: Runtime):
    if not _is_admin(rt, query.from_user.id):
        return
    await rt.db.set_setting("sys_channel", "")
    await query.answer("Канал убран", show_alert=True)
    text, kb = await _sysch_text_and_kb(rt)
    try:
        await query.message.edit_text(text, reply_markup=kb)
    except Exception:
        pass
